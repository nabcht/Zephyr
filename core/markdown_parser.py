"""MarkdownParser — extract text from PDF, Excel, and plain-text files as Markdown.

Preserves table structures by converting them to Markdown tables so LLMs can
reason about relational data accurately.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

log = logging.getLogger("uzephyr.markdown_parser")

# Extensions handled by each parser
_TEXT_EXTENSIONS: set[str] = {
    ".txt", ".md", ".py", ".json", ".csv", ".log", ".yaml", ".yml",
    ".toml", ".cfg", ".ini", ".html", ".xml", ".rst",
}
_PDF_EXTENSIONS: set[str] = {".pdf"}
_EXCEL_EXTENSIONS: set[str] = {".xlsx", ".xls"}
_DOCX_EXTENSIONS: set[str] = {".docx"}

# Maximum file size we will attempt to parse (10 MB)
_MAX_FILE_SIZE: int = 10 * 1024 * 1024


# ── Public API ────────────────────────────────────────────────────────────────

def parse_file(path: Path) -> list[dict[str, Any]]:
    """Parse a file into a list of content blocks.

    Each block is a dict:
        {"text": str, "source": str, "page": int | None}
    """
    if not path.is_file():
        return []
    if path.stat().st_size > _MAX_FILE_SIZE:
        log.debug("Skipping oversized file: %s", path)
        return []

    suffix = path.suffix.lower()
    try:
        if suffix in _PDF_EXTENSIONS:
            return _parse_pdf(path)
        if suffix in _EXCEL_EXTENSIONS:
            return _parse_excel(path)
        if suffix in _DOCX_EXTENSIONS:
            return _parse_docx(path)
        if suffix in _TEXT_EXTENSIONS:
            return _parse_text(path)
    except Exception as exc:
        log.warning("Failed to parse %s: %s", path, exc)
    return []


def chunk_text(
    text: str,
    chunk_words: int = 500,
    overlap_words: int = 50,
) -> list[str]:
    """Split *text* into overlapping word-level chunks."""
    words = text.split()
    if not words:
        return []
    chunks: list[str] = []
    start = 0
    while start < len(words):
        end = start + chunk_words
        chunks.append(" ".join(words[start:end]))
        start += chunk_words - overlap_words
    return chunks


# ── Private parsers ───────────────────────────────────────────────────────────

def _parse_pdf(path: Path) -> list[dict[str, Any]]:
    """Convert PDF pages to Markdown using PyMuPDF4LLM."""
    import pymupdf4llm  # lazy import — only needed when PDFs are found

    md_text = pymupdf4llm.to_markdown(str(path))
    if not md_text:
        return []
    # pymupdf4llm produces one big markdown string; split on page markers
    # (it inserts "-----" or form-feed markers between pages)
    pages = md_text.split("\n-----\n") if "\n-----\n" in md_text else [md_text]
    blocks: list[dict[str, Any]] = []
    for idx, page_text in enumerate(pages, 1):
        cleaned = page_text.strip()
        if cleaned:
            blocks.append({
                "text": cleaned,
                "source": path.name,
                "page": idx,
            })
    return blocks


def _parse_excel(path: Path) -> list[dict[str, Any]]:
    """Convert each Excel sheet into a Markdown table."""
    from openpyxl import load_workbook

    wb = load_workbook(str(path), read_only=True, data_only=True)
    blocks: list[dict[str, Any]] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])

        if not rows:
            continue

        # Build Markdown table
        header = rows[0]
        md_lines: list[str] = []
        md_lines.append("| " + " | ".join(header) + " |")
        md_lines.append("| " + " | ".join("---" for _ in header) + " |")
        for data_row in rows[1:]:
            # Pad or trim to header length
            padded = data_row + [""] * (len(header) - len(data_row))
            md_lines.append("| " + " | ".join(padded[: len(header)]) + " |")

        md_table = "\n".join(md_lines)
        blocks.append({
            "text": f"## Sheet: {sheet_name}\n\n{md_table}",
            "source": path.name,
            "page": None,
        })

    wb.close()
    return blocks


def _parse_text(path: Path) -> list[dict[str, Any]]:
    """Read a plain-text file and return it as a single block."""
    try:
        text = path.read_text(encoding="utf-8", errors="ignore")
    except (PermissionError, OSError):
        return []
    if not text.strip():
        return []
    return [{"text": text, "source": path.name, "page": None}]

def _parse_docx(path: Path) -> list[dict[str, Any]]:
    """Convert Word documents to Markdown text blocks."""
    try:
        import docx
    except ImportError:
        log.warning("python-docx not installed. Run 'pip install python-docx' for .docx support.")
        return []

    try:
        doc = docx.Document(path)
        paragraphs =[p.text.strip() for p in doc.paragraphs if p.text.strip()]
        
        if not paragraphs:
            return []
            
        text = "\n\n".join(paragraphs)
        return[{"text": text, "source": path.name, "page": None}]
    except Exception as exc:
        log.warning("Failed to parse DOCX %s: %s", path.name, exc)
        return[]